import hashlib
import os
import base64
import openpyxl

def generate_salt():
    """Generate a random salt."""
    return base64.urlsafe_b64encode(os.urandom(16)).decode()

def hash_with_salt(value, salt):
    """Hash the value with a given salt."""
    return hashlib.sha256((salt + value).encode()).hexdigest()

def save_to_excel(file_path, school_code, school_name, salt, hashed_value):
    """Save the original school code, school name, salt, and hashed value to an Excel file."""
    # Load existing workbook or create a new one if the file doesn't exist
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    # Select the active sheet or create a new sheet
    sheet = workbook.active
    sheet.title = "School Codes"

    # If the sheet is empty, add headers
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
        sheet.append(["School Code", "School Name", "Salt", "Salted Hash"])

    # Find the first empty row
    first_empty_row = sheet.max_row + 1

    # Append the new data at the first empty row
    sheet.append([school_code, school_name, salt, hashed_value])

    # Save the workbook
    workbook.save(file_path)
    print(f"Data for {school_code} saved to {file_path}")

# Dictionary of school codes and names
school_codes_names = {
    "XXX": " "
}

# Excel file path
excel_file_path = "Hash&salt_School_Code.xlsx"

# Process each school code, generate salt and hash, and save to Excel
for code, name in school_codes_names.items():
    salt = generate_salt()
    hashed_value = hash_with_salt(code, salt)
    save_to_excel(excel_file_path, code, name, salt, hashed_value)

print("All school codes processed and saved.")
